import openpyxl as xl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import date
from copy import copy

# column names CANNOT INCLUDE WHITESPACES
STOCK_KEY = '상품코드'
STOCK_LABELS = ['한글상품명', '색상', '옵션']
ORDER_KEY = '상품주문번호'
BUYING_LABELS = [
    '재고', '발송기한', '상품주문번호', '수령자', '주문일자', '카운터', '주문상태', '배송GBP', '이미지', '수량', '옵션', '상품코드', '원주문', '결제금액',
    # optional
    '주문번호', '주문메모', '배송메모', '상담메모', '원산지', '주문자ID', '전화번호', '핸드폰번호', '상품금액', '#', '주소', '우편번호', '개인통관부호', '배송메시지', '브랜드', '품목', '일반품목'
]

class CellFormat():
    def __init__(self, font, fill):
        self.font = copy(font)
        self.fill = copy(fill)

LABEL_FORMAT = CellFormat(Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="000000"))
DEFAULT_FORMAT = CellFormat(Font(), PatternFill())


class StockInfo():

    def __init__(self, stock_file_path, order_file_path):
        self.stock_file_path = stock_file_path
        self.order_file_path = order_file_path

        # 재고 parsed
        self.stock_data = dict() # { code : [{'상품명': xxx, '색상': xxx, '옵션': xxx, 'sheet' : xxx}, ...] }

        # 오더 parsed - orders without stock only
        self.rows_with_stock = set()
        self.buying_data = [] # [ { label : (raw_value, style) , ...}, ... ]
        self.buying_col_dimensions = []
        self.buying_row_dimensions = []


    # public methods
    def execute(self):
        stock_workbook = xl.load_workbook(self.stock_file_path)
        order_workbook = xl.load_workbook(self.order_file_path)

        self.iterate_stock(stock_workbook)
        self.iterate_order(order_workbook)

        self.leave_specific_sheet(order_workbook, '윈런던')
        self.highlight_sheet(order_workbook['윈런던'], self.rows_with_stock)

        order_workbook.save(f'{self.get_mmdd()} 새주문.xlsx')

        self.make_buying_list()        
    

    # private methods
    def iterate_stock(self, stock_workbook):
        for sheet in stock_workbook.sheetnames:
            table = stock_workbook[sheet]
            
            row_count = 0
            labels = []
            key_index = 0
            val_indices = []
            for row in table.iter_rows(max_col=15, values_only=True):
                row_count += 1

                if row_count == 1:
                    labels = [str(cell).replace(' ', '') for cell in row]
                    key_index = labels.index(STOCK_KEY)
                    val_indices = [labels.index(col) for col in STOCK_LABELS]
                    continue
                
                # build value dictionary
                val = {}
                for i in val_indices:
                    val[labels[i]] = str(row[i])
                val['sheet'] = sheet

                # store it in final dictionary
                key = str(row[key_index]).strip()
                if key not in self.stock_data:
                    self.stock_data[key] = [val]
                else:
                    self.stock_data[key].append(val)


    def iterate_order(self, order_workbook):
        order_worksheet = order_workbook['윈런던']

        row_count = 0
        labels = []
        val_indices = []
        for row_raw in order_worksheet.iter_rows():
            row_count += 1

            row = [ c.value for c in row_raw]
            try:
                if row_count == 1:
                    labels = [str(cell).replace(' ', '') for cell in row]
                    val_indices = []
                    buying_data_labels_row = []
                    for label in BUYING_LABELS:
                        buying_data_labels_row.append((label, LABEL_FORMAT))
                        if label == '이미지':
                            continue
                        val_indices.append(labels.index(label))
                    self.buying_data.append(buying_data_labels_row)
                    continue
                
                if row[0] == None:
                    continue

                if self.check_stock_for_order(self.stock_data, labels, row): # new order
                    self.rows_with_stock.add(row_count)

                else: # buying list
                    row_data = self.build_buying_row_data(order_worksheet, val_indices, row_raw)
                    self.buying_data.append(row_data)

            except Exception as e:
                raise Exception(f"Error: row {row_count} \n {row} \n {str(e)}")
            

    def check_stock_for_order(self, stock_data, order_labels, order_row):
        def get_val(label):
            index = order_labels.index(label)
            return str(order_row[index]).strip()

        if get_val('카운터') == '새주문' and (get_val('상품코드') in stock_data): # and that UI check!
            return True
        return False
    

    def build_buying_row_data(self, worksheet, val_indices, row_raw):
        row_data = []
        column_count = 0
        for i in val_indices:
            if column_count == 8:
                # add column for images
                row_data.append((' ', DEFAULT_FORMAT))
                self.buying_col_dimensions.append(26.5)
            value = row_raw[i].value
            format = CellFormat(row_raw[i].font, row_raw[i].fill)
            row_data.append((value, format))
            self.buying_col_dimensions.append(worksheet.column_dimensions[get_column_letter(i+1)].width)
            column_count += 1
        return row_data
            

    def make_buying_list(self):
        buying_list_workbook = xl.Workbook(write_only=True)
        buying_list_worksheet = buying_list_workbook.create_sheet('바잉리스트')

        # set dimensions - this MUST be done before adding cells
        for i in range(0, len(BUYING_LABELS)):
            col_i = i+1
            width = self.buying_col_dimensions[i]
            buying_list_worksheet.column_dimensions[get_column_letter(col_i)].width = width

        # append cells
        for row_data in self.buying_data:
            row = []
            for cell_data in row_data:
                cell = xl.cell.cell.Cell(buying_list_worksheet, value=cell_data[0])
                cell.font = cell_data[1].font
                cell.fill = cell_data[1].fill
                row.append(cell)
            buying_list_worksheet.append(row)

        buying_list_workbook.save(f'{self.get_mmdd()} 바잉리스트.xlsx')


    def highlight_sheet(self, worksheet, rows_to_highlight):
        fill_obj = PatternFill("solid", fgColor="FFFF00")
        row_count=0
        for row in worksheet.iter_rows(max_col=34):
            row_count += 1
            if row_count in rows_to_highlight:
                for cell in row:
                    cell.fill = fill_obj
            
    
    def get_mmdd(self):
        today = date.today()
        return today.strftime("%m%d")
        
    
    def leave_specific_sheet(self, workbook, worksheet_name):
        for sheet_name in workbook.sheetnames:
            if sheet_name != worksheet_name:
                sheet_temp = workbook[sheet_name]
                workbook.remove(sheet_temp)



stock = 'C:\\Users\\1004w\\Downloads\\재고문서 02-06 (마감) 백업본.xlsx'
order = 'C:\\Users\\1004w\\Downloads\\오더리스트 2023-02-06.xlsx'
test = StockInfo(stock, order)
test.execute()
#test.get_new_order_existing_stocks()
